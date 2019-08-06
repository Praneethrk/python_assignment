from openpyxl import Workbook
import openpyxl as xl

mark_map = {"S+":10,"S":9,"A":8,"B":7,"C":6,"D":5,"E":4,"F":0}

class Student:
    def __init__(self, usn,name):
        self.usn = usn
        self.name = name
        self.subjects = []

    def show_sgpa_info(self):
        g_c = 0
        s_c = 0
        for s in self.subjects:
            g_c += s["C"] * mark_map[s["G"]]
            s_c += s["C"]
        si = g_c / s_c
        print(f"{self.name} , {si}")
        

wb =Workbook()
wb = xl.load_workbook("./data/Result.xlsx")
sheet = wb.active
students = []
for row in sheet.iter_rows(min_row=3,min_col = 2,max_col = 13):
   if row:
       data = [c.value for c in row]
       stud_info = data[:2]
       sub_1 = data[5:7]
       sub_2  = data[-2:]
       student = Student(*stud_info)
       student.subjects.append({"C":sub_1[0],"G":sub_1[1]})
       student.subjects.append({"C":sub_2[0],"G":sub_2[1]})
       students.append(student)

for s in students:
    s.show_sgpa_info()

    